# region Librerías necesarias
import pandas as pd
import time
import requests
import logging
import threading
from datetime import datetime
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from itertools import product
from bs4 import BeautifulSoup
from langdetect import detect, DetectorFactory
import json
import re
import io
import random
from tqdm import tqdm
from codecarbon import OfflineEmissionsTracker
import streamlit as st
import unicodedata
from collections import Counter, defaultdict
import spacy
from spacy.cli import download
import matplotlib.pyplot as plt
from sklearn.preprocessing import MinMaxScaler
import matplotlib.pyplot as plt
import numpy as np
from fpdf import FPDF
import tempfile
import os
from PyPDF2 import PdfReader, PdfWriter

# endregion
# region Configuracion General
# region Configuración general de URLs y parámetros de ejecución
CROSSREF_API = "https://api.crossref.org/works"
IEPNB_API = "https://iepnb.gob.es/api/catalogo/v_listapatronespecie_normas"
EXCEL_URL = "https://www.miteco.gob.es/content/dam/miteco/es/biodiversidad/servicios/banco-datos-naturaleza/recursos/listas/lista-patron-especies-silvestres-con-normativa.xlsx"
OUTPUT_FILE = "ROSAL_IA.xlsx"
OUTPUT_FILE_AUX = "ROSAL_IA_aux.xlsx"
DELAY_BETWEEN_REQUESTS = 4  # Segundos entre peticiones a la API
RETRY_BACKOFF = 30           # Tiempo de espera tras recibir código 429
MAX_RETRIES = 10             # Reintentos máximos por fallo
DetectorFactory.seed = 0  # Para resultados reproducibles con langdetect
# endregion
# region Configuración de tracker de emisiones
pue = 1.12
country_iso_code = 'ESP'
region='ESP'
cloud_provider='gcp'
cloud_region='europe-southwest1'
country_2letter_iso_code = 'ES'
measure_power_secs = 30
save_to_file=False

tracker = OfflineEmissionsTracker(
        country_iso_code=country_iso_code,
        region=region,
        cloud_provider=cloud_provider,
        cloud_region=cloud_region,
        country_2letter_iso_code=country_2letter_iso_code,
        measure_power_secs=measure_power_secs,
        pue=pue,
        save_to_file= save_to_file
        )
# endregion
# region Configuración de registro (log) a consola y archivo
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
log_filename = f"ROSALIA_FETCHER_LOG_{timestamp}.txt"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(message)s",
    handlers=[
        logging.FileHandler(log_filename, mode='w', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
log = logging.info  # Alias para usar el log como si fuera print()

# Variables de control globales
completed_species = 0             # Contador de especies procesadas
lock = threading.Lock()          # Bloqueo para manejo seguro entre hilos
total_requests = 0               # Total de peticiones realizadas a Semantic Scholar

# Carga de filtros dinámicos desde Excel oficial
response = requests.get(EXCEL_URL)
response.raise_for_status()
filter_df = pd.read_excel(io.BytesIO(response.content), sheet_name=1)
# endregion
# endregion
# region configuración de filtrado 
# Extracción de columnas filtrables desde el archivo
# Lista fija de columnas relevantes
FILTER_COLUMNS = [
    "WithoutAutorship", "kingdom", "phylum", "class", "order", "family", "genus", "subgenus",
    "specificepithet", "infraspecificepithet", "taxonRank", "ScientificNameAuthorship",
    "taxonRemarks", "Vernacular Name", "Origen", "Environment", "Grupo taxonómico",
    "ScientificName", "Nombre en normativa", "Normativa", "Categoría",
    "Observaciones población", "Ámbito normativa", "Año normativa"
]

# Estructura simple, usada por funciones existentes
FILTER_OPTIONS = {
    col: sorted(filter_df[col].dropna().unique().tolist())
    for col in FILTER_COLUMNS if col in filter_df.columns
}

# Mapeo directo de nombres de filtro y DOIs
FILTER_KEY_MAP = {key: key for key in FILTER_OPTIONS.keys()}

# endregion

# region --- FUNCIONES AUXILIARES --- #

# Función para obtener el número de especies por filtro
def n_species_by_filter(key):
    """
    Devuelve un listado con el número de especies para cada valor único de un filtro simple.

    Args:
        key (str): Clave del filtro.

    Returns:
        list of tuples: Lista de tuplas (valor, número de especies)
    """
    results = []
    if key in FILTER_OPTIONS:
        for value in FILTER_OPTIONS[key]:
            filters = {key: f"eq.{value}"}
            species = fetch_species_list(filters=filters)
            count = len(species)
            results.append((value, count))
    return results


# Función para analizar combinaciones de filtros
def n_species_in_n_filters(base_filter_key, base_filter_value, compare_filter_keys):
    log(f"\nAnalizando filtro compuesto '{base_filter_key}' = '{base_filter_value}'")
    filter_combos = [FILTER_OPTIONS[key] for key in compare_filter_keys if key in FILTER_OPTIONS]
    results = []

    for values in product(*filter_combos):
        filters = {base_filter_key: f"eq.{base_filter_value}"}
        filters.update({k: f"eq.{v}" for k, v in zip(compare_filter_keys, values)})
        species = fetch_species_list(filters=filters)
        count = len(species)
        if count > 0:
            combo_desc = ", ".join(f"{k}={v}" for k, v in zip(compare_filter_keys, values))
            results.append((combo_desc, count))

    if results:
        log("\nFiltro combinado:")
        for combo, count in sorted(results, key=lambda x: -x[1]):
            log(f"- {combo}: {count} especies")
    else:
        log("No hay especies encontradas con esta combinación de filtros.")

    return results

# Función para obtener el abstract de un artículo utilizando la API de Semantic Scholar
def fetch_abstract_from_semantic_scholar(doi, title):
    """
    Intenta obtener el abstract de un artículo utilizando la API de Semantic Scholar.

    Args:
        doi (str): DOI del artículo.
        title (str): Título del artículo.

    Returns:
        str or None: El abstract si se encuentra, de lo contrario None.
    """
    api_url = "https://api.semanticscholar.org/graph/v1/paper/search?query="

    if doi:
        query = f"DOI:{doi}"
    elif title:
        query = title
    else:
        return None

    response = requests.get(f"{api_url}{query}&fields=abstract")
    if response.status_code == 200:
        data = response.json().get("data", [])
        if data and "abstract" in data[0]:
            return data[0]["abstract"]

    return None

# Función para obtener los metadatos de un artículo utilizando su DOI
def fetch_article_by_doi(doi, species_name, use_web_abstract=True):
    """
    Obtiene los metadatos de un artículo utilizando su DOI desde la API de CrossRef.

    Args:
        doi (str): Identificador DOI del artículo.
        species_name (str): Nombre de la especie científica asociada al artículo.
        use_web_abstract (bool): Indica si debe intentar obtener el abstract desde la web del artículo si no está presente.

    Returns:
        dict or None: Un diccionario con los metadatos del artículo si se recupera correctamente, de lo contrario None.
    """
    url = f"{CROSSREF_API}/{doi}"
    retries = 0

    while retries < MAX_RETRIES:
        try:
            response = requests.get(url)
            if response.status_code == 200:
                item = response.json().get("message", {})
                abstract = item.get("abstract", "")

                # Solo intentamos obtener abstract de la web si se permite
                if not abstract and item.get("URL") and use_web_abstract:
                    abstract = fetch_abstract_from_web(item.get("URL"))

                return {
                    "scientific name": species_name,
                    "title": item.get("title", [""])[0],
                    "year": extract_year(item),
                    "authors": format_authors(item.get("author", [])),
                    "abstract": abstract,
                    "url": item.get("URL", ""),
                    "DOI": doi
                }

            elif response.status_code == 429:
                time.sleep(RETRY_BACKOFF)
                retries += 1
            else:
                retries += 1
                time.sleep(DELAY_BETWEEN_REQUESTS)
        except Exception as e:
            log(f"⚠️ Error al obtener metadatos por DOI {doi}: {str(e)}")
            retries += 1

    return None

# Función para obtener el abstract desde la web del artículo
def fetch_abstract_from_web(url):
    """
    Intenta obtener el abstract de un artículo directamente desde su página web.

    Args:
        url (str): URL del artículo.

    Returns:
        str: El texto del abstract si se encuentra, de lo contrario una cadena vacía.
    """
    try:
        response = requests.get(url, timeout=120, allow_redirects=True)
        response.raise_for_status()
        final_url = response.url  # URL final después de redirecciones
        soup = BeautifulSoup(response.content, "html.parser")

        # Verificar si es un artículo de ScienceDirect
        if "sciencedirect.com" in final_url:
            abstract_section = soup.find("div", class_="abstract author")
            if abstract_section:
                abstract_text = abstract_section.get_text(strip=True, separator=" ")
                if abstract_text:
                    log(f"✅ Abstract obtenido desde ScienceDirect para: {final_url}")
                    return abstract_text

        # Verificar si es un artículo de TandFOnline
        if "tandfonline.com" in final_url:
            scripts = soup.find_all('script', type='application/ld+json')
            for script in scripts:
                try:
                    data = json.loads(script.string)
                    if isinstance(data, list):
                        for entry in data:
                            if entry.get("@type") == "ScholarlyArticle" and "abstract" in entry:
                                abstract_text = entry["abstract"].strip()
                                log(f"✅ Abstract obtenido desde TandFOnline para: {final_url}")
                                return abstract_text
                except json.JSONDecodeError:
                    continue

        # Procedimiento general para otros sitios
        abstract_keywords = ["abstract", "resumen", "summary"]
        possible_titles = soup.find_all(['h2', 'h3', 'strong', 'span', 'div'], string=True)

        for title in possible_titles:
            text_title = title.get_text(strip=True).lower()
            if any(keyword in text_title for keyword in abstract_keywords):
                next_elem = title.find_next_sibling()
                if next_elem and 50 < len(next_elem.get_text(strip=True)) < 2000:
                    log(f"✅ Abstract obtenido desde la web para: {final_url}")
                    return next_elem.get_text(strip=True)

        possible_abstracts = soup.find_all(['p', 'div'], string=True)
        for elem in possible_abstracts:
            text = elem.get_text(strip=True)
            if any(keyword in text.lower() for keyword in abstract_keywords) or (50 < len(text) < 2000):
                log(f"✅ Abstract obtenido desde la web para: {final_url}")
                return text

    except Exception as e:
        log(f"⚠️ No se pudo obtener el abstract de {url}: {str(e)}")

    return ""

# Validar que el artículo pertenece a la especie correcta o al género (con opción de abreviatura en paso 3)
def validate_species(article, species_name, genus=None, allow_abbreviation=False):
    """
    Valida si un artículo está relacionado con una especie específica.

    Args:
        article (dict): Diccionario con los metadatos del artículo.
        species_name (str): Nombre completo de la especie científica a validar.
        genus (str, opcional): Género de la especie para validación adicional.
        allow_abbreviation (bool): Indica si se permiten abreviaturas para la validación.

    Returns:
        bool: True si el artículo está relacionado con la especie, False en caso contrario.
    """
    full_name = species_name.lower()
    text_fields = (article.get("title", "") + " " + article.get("abstract", "")).lower()

    if full_name in text_fields:
        return True

    if allow_abbreviation:
        abbrev_1 = f"{species_name.split()[0][0]}. {species_name.split()[1]}".lower()
        abbrev_2 = f"{species_name.split()[0][0]}.{species_name.split()[1]}".lower()
        if abbrev_1 in text_fields or abbrev_2 in text_fields:
            return True

    if genus and genus.lower() in text_fields:
        return True

    return False

# Extraer el año del artículo
def extract_year(item):
    """
    Extrae el año de publicación de un artículo desde su estructura de metadatos.

    Args:
        item (dict): Diccionario de metadatos del artículo (CrossRef).

    Returns:
        int or None: El año de publicación si se encuentra, de lo contrario None.
    """
    try:
        if "published-online" in item:
            return item["published-online"]["date-parts"][0][0]
        if "published-print" in item:
            return item["published-print"]["date-parts"][0][0]
        if "issued" in item:
            return item["issued"]["date-parts"][0][0]
    except:
        return None

# Formatear autores
def format_authors(authors_list):
    """
    Formatea la lista de autores de un artículo en una cadena de texto legible.

    Args:
        authors_list (list): Lista de diccionarios con los datos de los autores.

    Returns:
        str: Cadena de texto con los nombres de los autores, separados por comas.
    """
    if not authors_list:
        return "Desconocido"
    return ", ".join([f"{a.get('given', '')} {a.get('family', '')}".strip() for a in authors_list])

# Detectar idioma
def detect_language(text):
    try:
        return detect(text)
    except:
        return "unknown"

def extract_english_block(text, min_non_en_block=30):
    # Divide el texto en frases o párrafos
    blocks = re.split(r'(?<=[.!?])\s+|\n+', text)
    english_blocks = []
    for block in blocks:
        block = block.strip()
        if len(block) == 0:
            continue
        try:
            lang = detect(block)
        except:
            lang = "unknown"
        # Si el bloque es inglés, lo guardamos
        if lang == "en":
            english_blocks.append(block)
        # Si el bloque NO es inglés y es corto, también lo guardamos (por si es un nombre científico, etc.)
        elif len(block) < min_non_en_block:
            english_blocks.append(block)
    # Si hay bloques en inglés (o cortos no ingleses), los unimos y devolvemos solo eso
    if english_blocks:
        return " ".join(english_blocks)
    # Si no hay bloques válidos, devolvemos vacío
    return ""

# endregion

# region --- FUNCIONES PRINCIPALES --- #

# Función para imprimir ayuda de filtros
def help_filters(key=None, value=None, num_filters=1):
    """
    Proporciona ayuda sobre los filtros disponibles y sus valores posibles para la API.

    Args:
        key (str or list, opcional): Clave del filtro o lista de claves para combinaciones.
        value (str, opcional): Valor del filtro base para mostrar combinaciones o ejemplo de uso.
        num_filters (int): Número de filtros a combinar. Por defecto 1 (individual).

    Flujo:
        - Sin clave: lista todas las claves de filtro disponibles.
        - Con clave y num_filters=1: muestra valores posibles con número de especies.
        - Con clave y valor: muestra ejemplo de uso (y combinaciones si es lista con num_filters > 1).

    Ejemplos:
        help_filters()  # Lista todas las claves de filtro.
        help_filters("Categoría")  # Muestra valores y especies para "Categoría".
        help_filters("Categoría", "Anexo II. Lista de especies en peligro o amenazadas")
        help_filters(["Categoría", "Grupo taxonómico"], value="Anexo II. Lista de especies en peligro o amenazadas", num_filters=2)

    Returns:
        None
    """
    # Caso 1: sin argumentos → mostrar claves
    if key is None:
        log("🔑 Claves de filtro disponibles:")
        for k in FILTER_OPTIONS:
            log(f"- {k}")
        return

    # Caso 2: combinación de filtros, requiere value del primer filtro
    if isinstance(key, list) and num_filters > 1:
        if len(key) != num_filters:
            log(f"⚠️ Debes proporcionar exactamente {num_filters} claves para combinaciones.")
            return
        if not value:
            log("ℹ️ Para combinaciones, proporciona el valor base del primer filtro.")
            log(f"Ejemplo: help_filters({key}, value=\"valor_base\", num_filters={num_filters})")
            return
        base_key = key[0]
        log(f"\n🔍 Combinaciones para {base_key} = '{value}':")
        combos = n_species_in_n_filters(base_key, value, key[1:])
        for desc, count in combos:
            log(f"   ↪ {desc} → {count} especies")
        return

    # Caso 3: filtro simple, mostrar valores y número de especies
    if isinstance(key, str) and key in FILTER_OPTIONS:
        if value is None:
            log(f"\n🎯 Valores posibles para '{key}' con número de especies:")
            values_with_counts = n_species_by_filter(key)
            for v, count in values_with_counts:
                log(f"- {v} → {count} especies")
            return
        else:
            log("📘 Ejemplo de uso con ese filtro:")
            log(f"filters = {{\"{key}\": \"eq.{value}\"}}")
            return

    log("⚠️ Clave(s) de filtro no válida(s). Usa help_filters() para ver las disponibles.")

# Función para establecer la lista de especies filtradas
def fetch_species_list(limit=1000, offset=0, filters=None):
    """
    Obtiene una lista de especies desde la API de IEPNB, aplicando filtros opcionales.

    Args:
        limit (int, opcional): Número máximo de especies a recuperar por solicitud (paginación).
        offset (int, opcional): Desplazamiento inicial para la paginación.
        filters (dict, opcional): Diccionario de filtros para refinar la búsqueda. 
            Las claves deben coincidir con las definidas en FILTER_KEY_MAP o en la API directamente.
            Para mayor rapidez en la creación y uso de filtros usar help_filters().
            Se pueden combinar múltiples filtros en un solo diccionario.

    Flujo:
        - Realiza solicitudes paginadas a la API de IEPNB hasta que no haya más datos.
        - Aplica filtros opcionales a cada solicitud para refinar los resultados.
        - Registra las URLs de las solicitudes para diagnóstico y depuración.
        - Si se encuentra un error en la respuesta de la API, se detiene el proceso.
        - Acumula todas las especies recuperadas en una lista.

    Returns:
        list: Lista de especies obtenidas desde la API de IEPNB.
    """
    log("Descargando lista de especies desde la API IEPNB...")
    species_list = []
    while True:
        params = {"limit": limit, "offset": offset}
        if filters:
            for k, v in filters.items():
                actual_key = FILTER_KEY_MAP.get(k, k)
                params[actual_key] = v
        response = requests.get(IEPNB_API, params=params)
        log("URL de la petición: " + response.url)
        if response.status_code != 200:
            log(f"Error en la descarga: {response.status_code}")
            break
        data = response.json()
        if not data:
            break
        species_list.extend(data)
        offset += limit
    log(f"Total de especies recuperadas: {len(species_list)}")
    return species_list

# Función para buscar artículos científicos para una especie específica
def fetcher_cf(species_name, n_species):
    """
    Obtiene artículos científicos para una especie específica desde la API de CrossRef.

    Args:
        species_name (str): Nombre de la especie científica.
        n_species (int): Número total de especies a procesar, para ajustar el límite de búsqueda.

    Returns:
        list: Lista de artículos recuperados desde CrossRef.
    """
    total_available_rows = 9900
    rows_per_species = total_available_rows // n_species
    if rows_per_species > 350:
        rows_per_species = 350
    log(f"🔍 Buscando artículos para la especie: {species_name}... | Máx. artículos: {rows_per_species}")

    genus = species_name.split()[0]
    query_url = (
        f"{CROSSREF_API}?query.bibliographic=\"{genus}\""
        f"&filter=type:journal-article&rows={rows_per_species}&sort=issued&order=desc&select=DOI"
    )

    response = requests.get(query_url)
    response.raise_for_status()
    data = response.json().get("message", {}).get("items", [])

    log(f"🔍 Artículos recuperados de Crossref: {len(data)} para la especie {species_name}.")
    return data

# Función para procesar los artículos obtenidos
def fetcher_processor(data, species_name):
    """
    Procesa una lista de artículos obtenidos, verificando su abstract y clasificándolos.

    Args:
        data (list): Lista de artículos obtenidos desde la API.
        species_name (str): Nombre de la especie científica.

    Returns:
        list: Lista de artículos procesados con información sobre abstract y criterio.
    """
    articles = []
    processed_dois = set()

    for item in data:
        doi = item.get("DOI")
        if doi not in processed_dois:
            processed_dois.add(doi)
            article = fetch_article_by_doi(doi, species_name, use_web_abstract=True)

            if article:
                if not article.get("abstract") and article.get("url"):
                    abstract = fetch_abstract_from_web(article.get("url"))
                    if not abstract:
                        abstract = fetch_abstract_from_semantic_scholar(doi, article.get("title"))

                    if abstract:
                        article["abstract"] = abstract

                article["abs_pres"] = 1 if article.get("abstract") else 0

                if validate_species(article, species_name, allow_abbreviation=True):
                    article["criterio"] = "Exacto"
                else:
                    article["criterio"] = "Genus"

                articles.append(article)

    log(f"🔎 Completado: {species_name} | Artículos procesados: {len(articles)}")
    return articles

# Función para buscar y procesar artículos en paralelo
def fetcher_pipe(species_name, n_species):
    """
    Combina la obtención y procesamiento de artículos científicos para una especie.

    Args:
        species_name (str): Nombre de la especie científica.
        n_species (int): Número total de especies a procesar.

    Returns:
        list: Lista de artículos procesados con información sobre abstract y criterio.
    """
    data = fetcher_cf(species_name, n_species)
    articles = fetcher_processor(data, species_name)
    return articles

# Función para limpiar los abstracts
def abstract_cleaning(df):
    """
    Limpia y mejora los abstracts de un DataFrame de artículos.

    Args:
        df (pd.DataFrame): DataFrame que contiene los artículos y sus abstracts.

    Returns:
        pd.DataFrame: DataFrame con los abstracts mejorados y la columna abs_pres ajustada.
    """
    log(f"Mejorando output de abstracts...")

    # Limpiar etiquetas HTML, saltos de línea y símbolos matemáticos al inicio en abstracts (pueden tener contenido matemático luego)
    df['abstract'] = df['abstract'].fillna("").apply(lambda x: re.sub(r'^[-=+*/%<>^&|]+', '', re.sub(r'<.*?>|\n|\r', ' ', x)).strip())

    # Verificar si el abstract es similar al título y eliminarlo, el abstract no puede ser lo mismo que el título. Lo hacemos antes de otras limpiezas para facilitar la detección exacta
    df['abstract'] = df.apply(lambda row: "" if row['title'].strip().lower() in row['abstract'].lower() and len(row['abstract']) <= len(row['title']) + 20 else row['abstract'], axis=1)
    
    # Eliminar la palabra "abstract" en cualquier variación de mayúsculas/minúsculas
    df['abstract'] = df['abstract'].str.replace(r'abstract', '', case=False, regex=True)

    # Eliminar "summary" o "summary:" si es la primera palabra del abstract en cualquier variación de mayúsculas/minúsculas
    df['abstract'] = df['abstract'].str.replace(r'^(summary:?)(\s+)', '', case=False, regex=True)

    # Eliminar "article" o "article:" si es la primera palabra del abstract en cualquier variación de mayúsculas/minúsculas
    df['abstract'] = df['abstract'].str.replace(r'^(article:?)(\s+)', '', case=False, regex=True)

    # Eliminar espacios duplicados
    df['abstract'] = df['abstract'].str.replace(r'\s{2,}', ' ', regex=True).str.strip()
   
    # Vaciar abstract si no está en inglés. Puede que haya abstracts en otros idiomas, pero generamente están mezclados con otros idiomas y/o con otros abstracts, siendo esta parte con otros abstract dañina para la calidad del dato, ganamos robustez
    df['abstract'] = df['abstract'].apply(lambda x: x if x == "" or detect_language(x) == "en" else "")

    # Extraer solo el bloque en inglés si hay texto mixto. Si el inglés está mezclado con otros idiomas pero es dominante en inglés, pasa el filtro anterior (abstract en varios idiomas). Esto nos devolverá solo el bloque en inglés
    df['abstract'] = df['abstract'].apply(lambda x: extract_english_block(x) if x else x)

    # Vaciar abstract si la primera palabra de 'scientific name' no aparece en el abstract. Se perderán abstract pero se gana en calidad del dato, evitando abstracts emplazados erroneamente.
    df['abstract'] = df.apply(lambda row: row['abstract'] if row['abstract'] == "" or row['scientific name'].split()[0].lower() in row['abstract'].lower() else "",axis=1)

    # Eliminar abstracts que contienen el mensaje de compra completada o de artículos retirados
    df['abstract'] = df['abstract'].apply(lambda x: "" if "Your purchase has been completed" in x else x)

    df['abstract'] = df['abstract'].apply(lambda x: "" if "This retracts the article" in x else x)

    df['abstract'] = df.apply(lambda row: "" if re.match(r"^[\[\(\s]{0,2}retracted[\]\)\s]{0,2}", row['title'].strip(), re.IGNORECASE) else row['abstract'],axis=1)

    # Ajustar abs_pres a 0 si el abstract está vacío para facilitar procesamiento posterior
    df['abs_pres'] = df['abstract'].apply(lambda x: 0 if x == "" else 1)

    log(f"Abstracts mejorados.")
    return df

# Función para actualizar artículos de especies
def update_species_articles(filters=None, streamlit_mode=False):
    """
    Actualiza los artículos científicos para una lista de especies, procesándolos de manera paralela.
    Si se usa "_species" en filters, se filtra directamente sobre la lista devuelta de la API.

    Args:
        filters (dict, opcional): Diccionario de filtros. "_species" se usa para filtrar internamente.
        streamlit_mode (bool): Si True, muestra progreso en Streamlit.

    Returns:
        pd.DataFrame o None: DataFrame con resultados si hay datos, si no None.
    """
    filters = filters or {}
    # Si solo se usa _species, no hace falta llamar a fetch_species_list
    if "_species" in filters:
        species_names = list(dict.fromkeys(filters["_species"]))  # Mantiene el orden y unicidad
        species_list = [{"WithoutAutorship": s} for s in species_names]
    else:
        filtros_api = {k: v for k, v in filters.items() if k != "_species"}
        species_list = fetch_species_list(filters=filtros_api)

    total_species = len(species_list)
    all_results = []

    log(f"Iniciando búsqueda de artículos para {total_species} especies.")
    start_time = time.time()

    if streamlit_mode:
        progress_bar = st.progress(0)
        progress_text = st.empty()
    else:
        pbar = tqdm(total=total_species, desc="Procesando especies", unit="especies")

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {
            executor.submit(fetcher_pipe, species["WithoutAutorship"], total_species): species
            for species in species_list
        }

        completed = 0

        for future in as_completed(futures):
            species = futures[future]
            articles = future.result()
            all_results.extend(articles)

            exact_count = sum(1 for a in articles if a['criterio'] == 'Exacto')
            genus_count = sum(1 for a in articles if a['criterio'] == 'Genus')
            exact_with_abstract = sum(1 for a in articles if a['criterio'] == 'Exacto' and a['abs_pres'] == 1)
            genus_with_abstract = sum(1 for a in articles if a['criterio'] == 'Genus' and a['abs_pres'] == 1)

            log(f"🔎 Completado: {species['WithoutAutorship']} | Total: {len(articles)} | "
                f"Exacto: {exact_count} (Con abstract: {exact_with_abstract}, Sin: {exact_count - exact_with_abstract}) | "
                f"Genus: {genus_count} (Con abstract: {genus_with_abstract}, Sin: {genus_count - genus_with_abstract})")

            completed += 1
            elapsed_time = time.time() - start_time
            percent_complete = int((completed / total_species) * 100)

            if streamlit_mode:
                progress_bar.progress(percent_complete)
                progress_text.markdown(f"✅ Procesadas: {completed}/{total_species} especies ({percent_complete}%)")
            else:
                pbar.set_postfix({"Tiempo": f"{round(elapsed_time / 60, 2)} min"})
                pbar.update(1)

    if not streamlit_mode:
        pbar.close()

    if all_results:
        df = pd.DataFrame(all_results)
        df = df.sort_values(by=["scientific name", "year", "criterio"], ascending=[True, False, True])
        df = abstract_cleaning(df)
        df.to_excel(OUTPUT_FILE, index=False)

        wb = load_workbook(OUTPUT_FILE)
        wb.properties.keywords = f"Filtros usados: {filters}"
        wb.save(OUTPUT_FILE)
        wb.close()

        log(f"\n📁 Archivo generado: {OUTPUT_FILE}")
    else:
        log("\n🚫 No se encontraron artículos nuevos.")

    total_time = time.time() - start_time
    log(f"\n⏱️ Tiempo total: {round(total_time / 60, 2)} minutos")

    return df if all_results else None


# endregion

# region reporter

model_name = "en_core_web_lg"
log("Cargando modelo spaCy en_core_web_lg...")
try:
    nlp = spacy.load(model_name)
except OSError:
    log("Modelo en_core_web_lg no encontrado, descargando modelo...")
    download(model_name)
    nlp = spacy.load(model_name)

def clean_text(text):
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r'[\t\r\x0b\x0c]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def deduplicate_sentences(sentences):
    seen = set()
    result = []
    for sent in sentences:
        s = sent.lower()
        if s not in seen:
            seen.add(s)
            result.append(sent)
    return result

def summarize_with_spacy(abstracts, max_chars=3000):
    """
    Genera un resumen extractivo coherente a partir de múltiples abstracts científicos.
    Las frases seleccionadas se priorizan por relevancia semántica, pero se limita el número
    de frases por abstract para evitar sobre-representación y se ordenan según su aparición
    original para mantener coherencia narrativa.

    Args:
        abstracts (list of str): Lista de textos científicos.
        max_chars (int): Límite de caracteres del resumen.

    Returns:
        str: Resumen limpio y cohesivo.
    """
    from collections import defaultdict

    all_text = " ".join(abstracts)
    doc = nlp(all_text)

    # Extraer todas las frases con su índice de abstract de origen
    sentence_map = []  # (sentence_text, abstract_idx, position_in_text)
    abstract_offset = 0
    for idx, abs_text in enumerate(abstracts):
        abs_doc = nlp(abs_text)
        for sent in abs_doc.sents:
            sent_text = sent.text.strip()
            if len(sent_text) > 50:
                sentence_map.append((sent_text, idx, abstract_offset))
            abstract_offset += 1

    # Calcular frecuencia de palabras clave
    keywords = [t.lemma_.lower() for t in doc if t.pos_ in ["NOUN", "PROPN"] and not t.is_stop]
    freq = Counter(keywords)

    # Rankear frases por puntuación semántica
    ranked = sorted(
        sentence_map,
        key=lambda tup: sum(freq.get(w.lemma_.lower(), 0) for w in nlp(tup[0])),
        reverse=True
    )

    # Eliminar duplicados (por texto)
    seen = set()
    deduped = []
    for text, idx, pos in ranked:
        if text not in seen:
            deduped.append((text, idx, pos))
            seen.add(text)

    # Limitar el número de frases por abstract (máx 4 por abstract)
    abstract_sentence_counts = defaultdict(int)
    selected = []
    total_chars = 0

    for sentence, abs_idx, pos in deduped:
        if abstract_sentence_counts[abs_idx] >= 4:
            continue
        if total_chars + len(sentence) > max_chars:
            break
        selected.append((sentence, pos))
        abstract_sentence_counts[abs_idx] += 1
        total_chars += len(sentence)

    # Ordenar por posición original para mayor fluidez narrativa
    selected.sort(key=lambda tup: tup[1])
    final_sentences = [s for s, _ in selected]

    return clean_text(" ".join(final_sentences))

def generate_summary_for_species(df, especie, criterio="Exacto"):
    sub_df = df[
        (df["scientific name"] == especie) &
        (df["criterio"] == criterio) &
        (df["abs_pres"] == 1)
    ]
    abstracts = sub_df["abstract"].dropna().tolist()
    total = len(abstracts)

    if total == 0:
        return None

    use_n = total
    max_chars = 3000 if total <= 5 else 3000 + (total - 5) * 800

    if max_chars > 12000:
        st.warning(f"⚠️ La especie **{especie}** tiene {total} artículos ({criterio}).")
        opcion = st.radio(
            f"Resumen para {especie} excede el límite. ¿Qué hacer?",
            [f"Usar primeros 20 artículos",
             f"Incluir los {total} artículos",
             f"Indicar manualmente cuántos incluir"],
            key=f"opcion_{criterio}_{especie}"
        )
        if "20" in opcion:
            use_n = min(20, total)
        elif "Indicar" in opcion:
            use_n = st.number_input(
                f"Número de artículos para {especie} ({criterio}) (máx. {total})",
                min_value=1,
                max_value=total,
                step=1,
                key=f"custom_n_{criterio}_{especie}"
            )
        else:
            use_n = total

    abstracts = abstracts[:int(use_n)]
    n = len(abstracts)
    max_chars = 3000 if n <= 5 else min(12000, 3000 + (n - 5) * 800)

    full_text = " ".join(abstracts)
    doc = nlp(full_text)
    keywords = [t.lemma_.lower() for t in doc if t.pos_ in ["NOUN", "PROPN"] and not t.is_stop]
    top_keywords = [kw for kw, _ in Counter(keywords).most_common(10)]

    resumen = abstracts[0] if n == 1 else summarize_with_spacy(abstracts, max_chars=max_chars)
    referencias = sub_df.iloc[:int(use_n)][["scientific name", "title", "year", "authors", "url"]].to_dict("records")

    if criterio == "Genus":
        disclaimer = (
            "⚠️ *Este texto se ha generado con artículos científicos relacionados a nivel de género. "
            "Puede incluir especies distintas dentro del género, lo que reduce precisión.*\n\n"
        )
        resumen = disclaimer + resumen

    return {
        "resumen": resumen,
        "referencias": referencias,
        "palabras_clave": top_keywords,
        "num_abstracts": int(use_n)
    }

def generate_scientific_report_data(df_resultado):
    """
    Genera la estructura de datos para el informe científico a partir de un DataFrame
    de artículos. Permite seleccionar qué especies incluir (por criterio) y cuántos
    artículos usar, antes de generar los resúmenes. La interacción es especie por especie
    si el usuario elige selección manual.

    Args:
        df_resultado (pd.DataFrame): DataFrame con artículos procesados.

    Returns:
        dict: Diccionario con claves 'especificos', 'genericos' y 'graficas_calidad' según selección del usuario.
    """
    report_data = {}

    df_resultado['title'] = df_resultado['title'].fillna("").apply(clean_text)
    df_resultado['abstract'] = df_resultado['abstract'].fillna("").apply(clean_text)

    # === ESPECIES DISPONIBLES ===
    conteo_exactos = df_resultado[
        (df_resultado['criterio'] == 'Exacto') & (df_resultado['abs_pres'] == 1)
    ]['scientific name'].value_counts()

    conteo_genus = df_resultado[
        (df_resultado['criterio'] == 'Genus') & (df_resultado['abs_pres'] == 1)
    ]['scientific name'].value_counts()

    # === CONTROL ESPECÍFICOS ===
    if not conteo_exactos.empty:
        st.markdown("### 📌 Artículos específicos disponibles por especie:")
        for especie, count in conteo_exactos.items():
            st.markdown(f"- **{especie}** → {count} artículos con abstract")

    report_data["especificos"] = {}
    st.subheader("📄 ¿Deseas generar informes específicos (criterio Exacto)?")
    opcion_exacto = st.radio("Selecciona una opción:", ["Sí", "No", "Depende (Selección manual)"], key="radio_opcion_exacto")

    especies_exacto = sorted(conteo_exactos.index)

    if opcion_exacto == "Sí":
        for especie in especies_exacto:
            report_data['especificos'][especie] = generate_summary_for_species(df_resultado, especie, criterio="Exacto")

    elif opcion_exacto == "Depende (Selección manual)":
        seleccionadas = st.multiselect("Selecciona las especies para generar informe específico:", especies_exacto)
        for especie in seleccionadas:
            report_data['especificos'][especie] = generate_summary_for_species(df_resultado, especie, criterio="Exacto")

    # === CONTROL GENÉRICOS ===
    if not conteo_genus.empty:
        st.markdown("### 📌 Artículos genéricos disponibles por especie:")
        for especie, count in conteo_genus.items():
            st.markdown(f"- **{especie}** → {count} artículos con abstract")
    
    report_data["genericos"] = {}
    st.subheader("🧬 ¿Deseas generar informes genéricos (criterio Genus)?")
    opcion_genus = st.radio("Selecciona una opción:", ["Sí", "No", "Depende (Selección manual)"], key="radio_opcion_genus")

    especies_genus = sorted(conteo_genus.index)

    if opcion_genus == "Sí":
        for especie in especies_genus:
            report_data['genericos'][especie] = generate_summary_for_species(df_resultado, especie, criterio="Genus")

    elif opcion_genus == "Depende (Selección manual)":
        seleccionadas = st.multiselect("Selecciona las especies para generar informe genérico:", especies_genus)
        for especie in seleccionadas:
            report_data['genericos'][especie] = generate_summary_for_species(df_resultado, especie, criterio="Genus")
       
    # === GRÁFICAS POR ESPECIE ===
    report_data["graficas_calidad"] = None

    if st.checkbox("📈 ¿Deseas también incorporar gráficos sobre la calidad de los datos en el informe?"):
        df_indicadores = generate_quality_indicators(df_resultado)
        publication_history = generate_publication_history_charts(df_resultado)
        radar_charts = {}
        for especie in df_resultado['scientific name'].unique():
            radar_charts[especie] = plot_radar_chart(df_indicadores, especie)
        report_data["graficas_calidad"] = {
            "indicadores": df_indicadores,
            "radar_charts": radar_charts,
            "publication_history": publication_history
        }
        st.markdown("""
        ### 📊 Indicadores de calidad de datos científicos
        Las siguientes métricas se han calculado por especie para evaluar la calidad de la información:
        - **Recencia**: Año promedio de publicación de los artículos.
        - **Cantidad**: Número total de artículos recuperados.
        - **Precisión**: Proporción de artículos 'Exacto' entre todos los artículos con abstract.
        - **Cobertura**: Proporción de artículos que contienen abstract.
        - **Diversidad temporal**: Dispersión en años de publicación (desviación estándar).

        Además, se calcula un **Índice Global de Calidad** (IGC) como la suma normalizada de los anteriores.
        """)

    return report_data

# region pdf

def move_index_to_second_page(pdf_path, output_path):
    reader = PdfReader(pdf_path)
    writer = PdfWriter()

    # Asumiendo que el índice es la última página
    portada = reader.pages[0]
    indice = reader.pages[-1]
    resto = reader.pages[1:-1]

    writer.add_page(portada)
    writer.add_page(indice)
    for page in resto:
        writer.add_page(page)

    with open(output_path, "wb") as f_out:
        writer.write(f_out)

def generate_pdf_report(report_data, filtros_aplicados, timestamp, df_resultado):
  
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Validación de fuentes
    font_dir = os.path.dirname(__file__)
    fuentes = {
        "": "DejaVuSans.ttf",
        "B": "DejaVuSans-Bold.ttf",
        "I": "DejaVuSans-Oblique.ttf",
        "BI": "DejaVuSans-BoldOblique.ttf"
    }

    for estilo, archivo in fuentes.items():
        ruta_fuente = os.path.join(font_dir, archivo)
        if not os.path.isfile(ruta_fuente):
            raise FileNotFoundError(f"Fuente no encontrada: {ruta_fuente}")
        pdf.add_font("DejaVu", estilo, ruta_fuente, uni=True)

    pdf.set_font("DejaVu", '', 12)
    # Footer mejorado
    pdf.footer = lambda: (
        None if pdf.page_no() in {1, 2} else (
            pdf.set_y(-15),
            pdf.set_font("DejaVu", '', 8),
            pdf.cell(
                0, 10,
                f"{' Repository Of Scientific Articles on Listed species - Informe científico generado   ' + str(pdf.page_no()) if pdf.page_no() % 2 == 0 else str(pdf.page_no()) + '   Repository Of Scientific Articles on Listed species - Informe científico generado'}",
                align='L' if pdf.page_no() % 2 == 0 else 'R'
            )
        )
    )
    
    # Portada
    pdf.add_page()
    pdf.set_font("DejaVu", 'B', 18)
    pdf.multi_cell(0, 10, "ROSAL.IA - Repository Of Scientific Articles on Listed species - Informe Científico Automatizado", align='C')
    pdf.ln(10)
    pdf.set_font("DejaVu", '', 12)
    pdf.multi_cell(0, 8, f"Fecha de generación: {timestamp}")
    pdf.ln(5)
    pdf.multi_cell(0, 8, "Filtros aplicados:")
    for f in filtros_aplicados:
        pdf.multi_cell(0, 8, f"- {f['tipo'].capitalize()}: {f['clave']} = {f['valor']}")
    total_especies = df_resultado['scientific name'].nunique()
    pdf.ln(5)
    pdf.multi_cell(0, 8, f"Número total de especies analizadas: {total_especies}")

    # Por especie
    index_entries = []
    especies = sorted(set(list(report_data.get("especificos", {}).keys()) + list(report_data.get("genericos", {}).keys())))
    for especie in especies:
        pdf.add_page()
        start_page = pdf.page_no()
        index_entries.append((especie, start_page))
        pdf.set_font("DejaVu", 'B', 14)
        pdf.multi_cell(0, 10, especie)
        pdf.ln(3)

        # Resumen específico (inglés)
        datos = report_data.get("especificos", {}).get(especie)
        if datos:
            pdf.set_font("DejaVu", 'I', 11)
            pdf.multi_cell(0, 8, "Resumen específico:")
            pdf.ln(2)
            pdf.set_font("DejaVu", '', 11)
            pdf.multi_cell(0, 8, datos.get("resumen", ""))
            pdf.ln(3)            
            pdf.set_font("DejaVu", 'I', 11)
            pdf.multi_cell(0, 8, "Referencias específicas:")
            pdf.ln(2)
            pdf.set_font("DejaVu", '', 10)
            for ref in datos["referencias"]:
                ref_txt = f"- {ref['title']} ({ref['year']}) - {ref['authors']} {ref['url']}"
                pdf.multi_cell(0, 8, ref_txt)
            pdf.ln(5)

        # Resumen genérico (inglés)
        datos_gen = report_data.get("genericos", {}).get(especie)
        if datos_gen:
            pdf.set_font("DejaVu", 'I', 11)
            pdf.multi_cell(0, 8, "Resumen genérico:")
            pdf.ln(2)
            pdf.set_font("DejaVu", '', 11)
            pdf.multi_cell(0, 8, datos_gen.get("resumen", ""))
            pdf.ln(3)            
            pdf.set_font("DejaVu", 'I', 11)
            pdf.multi_cell(0, 8, "Referencias genéricas:")
            pdf.ln(2)
            pdf.set_font("DejaVu", '', 10)
            for ref in datos_gen["referencias"]:
                ref_txt = f"- {ref['title']} ({ref['year']}) - {ref['authors']} {ref['url']}"
                pdf.multi_cell(0, 8, ref_txt)
            pdf.ln(5)

        # Gráficas de calidad por especie EN PÁGINA NUEVA
        radar_fig = report_data.get("graficas_calidad", {}).get("radar_charts", {}).get(especie)
        pub_hist_dict = report_data.get("graficas_calidad", {}).get("publication_history", {}).get(especie, {})

        if radar_fig or pub_hist_dict:
            pdf.add_page()
            pdf.set_font("DejaVu", 'B', 12)
            pdf.cell(0, 10, "Gráficas de calidad", ln=True)
            if radar_fig:
                tmpfile = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                radar_fig.savefig(tmpfile.name)
                pdf.image(tmpfile.name, x=10, y=None, w=90)
            x_offset = 110
            for criterio, fig in pub_hist_dict.items():
                tmpfile = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                fig.savefig(tmpfile.name)
                pdf.image(tmpfile.name, x=x_offset, y=None, w=90)
                x_offset += 100

    # Página de índices de calidad y tabla general (DESPUÉS de las gráficas y ANTES de referencias)
    pdf.add_page()
    pdf.set_font("DejaVu", 'B', 16)
    pdf.cell(0, 12, "Índices de calidad", ln=True, align='C')
    pdf.ln(4)
    pdf.set_font("DejaVu", '', 12)
    pdf.multi_cell(0, 8, """📊 Indicadores de calidad de datos científicos
Las siguientes métricas se han calculado por especie para evaluar la calidad de la información:
- **Recencia**: Año promedio de publicación de los artículos.
- **Cantidad**: Número total de artículos recuperados.
- **Precisión**: Proporción de artículos 'Exacto' entre todos los artículos con abstract.
- **Cobertura**: Proporción de artículos que contienen abstract.
- **Diversidad temporal**: Dispersión en años de publicación (desviación estándar).

Además, se calcula un **Índice Global de Calidad** (IGC) como la suma normalizada de los anteriores.
""")
    pdf.ln(4)
    
    # Página de referencias a APIs (siempre al final)
    pdf.add_page()
    pdf.set_font("DejaVu", 'B', 14)
    pdf.cell(0, 10, "Referencias: APIs y librerías", ln=True)
    pdf.set_font("DejaVu", '', 11)
    refs = [
        "- Crossref. (s.f.). Crossref REST API. https://api.crossref.org",
        "- Allen Institute for AI. Semantic Scholar Academic Graph API. s.f. Web. https://api.semanticscholar.org/",
        "- Ministerio para la Transición Ecológica y el Reto Demográfico. Servicios interoperables del Inventario Español del Patrimonio Natural y la Biodiversidad. s.f. Web. https://iepnb.gob.es/en/resources/interoperable-services",
        "- pandas, numpy, matplotlib, fpdf, spaCy, scikit-learn, streamlit"
    ]
    for ref in refs:
        pdf.multi_cell(0, 8, ref)

    # Índice (al final, sin footer)
    pdf.footer = lambda: None
    pdf.add_page()
    pdf.set_font("DejaVu", 'B', 14)
    pdf.cell(0, 10, "Índice", ln=True)
    pdf.set_font("DejaVu", '', 12)
    pdf.cell(0, 10, "- Resúmenes por especie:", ln=True)
    for especie, page in index_entries:
        pdf.cell(0, 10, f"    {especie} .......... {page}", ln=True)

    # Guardar PDF temporal
    tmp_pdf_path = f"Informe_ROSALIA_{timestamp}_tmp.pdf"
    pdf.output(tmp_pdf_path)

    # Reordenar para que el índice sea la segunda página
    final_pdf_path = f"Informe_ROSALIA_{timestamp}.pdf"
    move_index_to_second_page(tmp_pdf_path, final_pdf_path)

    # Limpieza del temporal
    try:
        os.remove(tmp_pdf_path)
    except Exception:
        pass

    return final_pdf_path

# endregion

# region graphics

def generate_publication_history_charts(df_resultado):
    """
    Genera gráficos de líneas con puntos para el histórico de publicaciones por especie y criterio.
    Devuelve un diccionario: {especie: {"Exacto": fig, "Genus": fig}}
    """
    charts = defaultdict(dict)
    especies = df_resultado["scientific name"].unique()

    for especie in especies:
        for criterio in ["Exacto", "Genus"]:
            sub = df_resultado[
                (df_resultado["scientific name"] == especie) &
                (df_resultado["criterio"] == criterio) &
                (df_resultado["year"].notnull())
            ]
            if sub.empty:
                continue

            counts = sub.groupby("year").size().sort_index()
            fig, ax = plt.subplots(figsize=(6, 4))
            ax.plot(counts.index, counts.values, marker="o")
            ax.set_title(f"{especie} - {criterio}")
            ax.set_xlabel("Año")
            ax.set_ylabel("Nº de artículos")
            ax.grid(True)
            charts[especie][criterio] = fig

    return charts

def generate_quality_indicators(df_resultado):
    """
    Calcula métricas de calidad de los datos científicos por especie para graficar en radar chart.

    Returns:
        pd.DataFrame: DataFrame con las métricas normalizadas por especie.
    """
    
    quality_data = []
    species_list = df_resultado['scientific name'].unique()

    for especie in species_list:
        subset = df_resultado[df_resultado['scientific name'] == especie]
        total = len(subset)
        con_abstract = subset['abs_pres'].sum()
        exacto = subset[(subset['criterio'] == 'Exacto') & (subset['abs_pres'] == 1)]
        genus = subset[(subset['criterio'] == 'Genus') & (subset['abs_pres'] == 1)]

        recencia = subset['year'].dropna()
        recencia_score = recencia.mean() if not recencia.empty else 0
        diversidad = recencia.std() if not recencia.empty else 0

        calidad = {
            'Especie': especie,
            'Recencia': recencia_score,
            'Cantidad': total,
            'Precisión': len(exacto) / (len(exacto) + len(genus)) if (len(exacto) + len(genus)) > 0 else 0,
            'Cobertura': con_abstract / total if total > 0 else 0,
            'Diversidad temporal': diversidad
        }
        calidad['Índice Global'] = sum([
            calidad['Recencia'],
            calidad['Cantidad'],
            calidad['Precisión'],
            calidad['Cobertura'],
            calidad['Diversidad temporal']
        ])  # se normaliza luego

        quality_data.append(calidad)

    df_quality = pd.DataFrame(quality_data)
    df_quality.set_index('Especie', inplace=True)

    scaler = MinMaxScaler()
    df_quality[df_quality.columns] = scaler.fit_transform(df_quality[df_quality.columns])

    return df_quality.T.reset_index().rename(columns={'index': 'Feature'})

def plot_radar_chart(df_quality, especie, save_path=None):
    data = df_quality.copy()
    data = data.loc[(data.iloc[:, 1:] != 0).any(axis=1)]  # eliminar columnas sin variación

    features = data['Feature'].tolist()
    if especie not in data.columns:
        return None
    values = data[especie].values
    num_vars = len(features)
    angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
    angles += angles[:1]  # cerrar gráfico

    # Obtener el IGC para la especie
    igc = None
    if "Índice Global" in features:
        igc_idx = features.index("Índice Global")
        igc = values[igc_idx]
    igc_str = f" (IGC = {igc:.2f})" if igc is not None else ""

    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    values = np.concatenate((values, [values[0]]))  # cerrar gráfico
    ax.plot(angles, values, linewidth=2, label=especie)
    ax.fill(angles, values, alpha=0.25)
    ax.set_yticklabels([])
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(features, fontsize=10)
    plt.title(f"{especie}{igc_str}", fontsize=13)  # Título con IGC
    plt.legend(loc='upper right', bbox_to_anchor=(1.1, 1))
    if save_path:
        fig.savefig(save_path)
    plt.close(fig)
    return fig

# endregion

# endregion

# region streamlit

st.set_page_config(page_title="ROSAL.IA Chatbot Reporter", layout="centered")

# Estructura enriquecida, usada para interfaz Streamlit (con especies y conteo)

# Lista de posibles mensajes de carga mientras se prepara la estructura enriquecida, para no tener el mensaje de running function por la carga de caché
loading_messages = [
    "🐝 Polinizando filtros...",
    "🦎 Cazando datos en la maleza...",
    "🌱 Germinando opciones de búsqueda...",
    "🐾 Rastreando huellas de especies...",
    "🪲 Clasificando biodiversidad...",
    "🧬 Secuenciando especies...",
    "🌍 Explorando el ecosistema digital...",
    "🦋 Atrayendo especies con colores vivos..."
]

@st.cache_data(show_spinner=random.choice(loading_messages))
def get_enriched_filter_options():
    enriched = {}
    for col in FILTER_COLUMNS:
        if col in filter_df.columns:
            enriched[col] = {}
            for val in sorted(filter_df[col].dropna().unique()):
                subset = filter_df[filter_df[col] == val]
                species = sorted(subset["WithoutAutorship"].dropna().unique().tolist())
                enriched[col][val] = {
                    "count": len(species),
                    "species": species
                }
    return enriched

FILTER_ENRICHED_OPTIONS = get_enriched_filter_options()


# --- 1. Bienvenida anclada ---
st.markdown("""
# 🌿 Bienvenid@ al Science Desk de **ROSAL.IA**
📚 *Repository Of Scientific Articles on Listed species*

A través de preguntas, sacaremos el informe basado en abstracts de artículos científicos recientes sobre la lista de especies catalogadas de tu interés.

**IMPORTANTE**: Este generador de informes científicos está hecho con base NLP, no LLM, por lo que las preguntas y respuestas son guiadas para su correcto funcionamiento. Al ser NLP y no LLM con RAG, evitamos problemas de alucinación y sesgo en los resultados, manteniéndonos fieles a lo escrito en los artículos.

También se fundamenta en la integración de APIs del catálogo de especies del [IEPNB](https://iepnb.gob.es/recursos/servicios-interoperables/api-catalogo), búsqueda de artículos en [Crossref](https://api.crossref.org) y [Semantic Scholar](https://api.semanticscholar.org).

🧠 Proyecto dentro del **Hackathon del Programa Nacional de Algoritmos Verdes (PNAV)**, en colaboración con el **Ministerio de Transición Ecológica (MITECO), TRAGSA y Accenture**. Desarrollado por el grupo [AI.IDEA](https://github.com/AEDI-IA/Ai.dea?tab=readme-ov-file#aiidea).

---
""")

# --- 2. Selección dinámica y combinación de filtros con selección manual de especies --- #
st.markdown("---")
st.markdown("### 🧪 Selección y combinación de filtros sobre especies")
st.markdown("Si quieres filtrar por más de un criterio, puedes hacerlo. Puedes combinar filtros(Intersección) para acotar aún más la búsqueda o añadir filtros(unión) si te interesa ver varios juntos. **Elige el modo del filtro que quieres aplicar y selecciona la categoría y el valor. Puedes añadir tantos filtros como quieras, pero ten en cuenta que si usas demasiados filtros, puede que no haya especies que cumplan todos los criterios.**")

if st.button("🔄 Empezar selección de filtros de cero"):
    st.session_state.filtros_aplicados = []
    st.session_state.especies_totales = []
    st.session_state.especies_seleccionadas_finales = []
    st.experimental_rerun()

if "filtros_aplicados" not in st.session_state:
    st.session_state.filtros_aplicados = []
if "especies_totales" not in st.session_state:
    st.session_state.especies_totales = []
if "especies_seleccionadas_finales" not in st.session_state:
    st.session_state.especies_seleccionadas_finales = []

# Selector de tipo de filtro
modo = st.radio("¿Qué quieres hacer ahora?", ["➕ Añadir filtro (Unión)", "🔗 Combinar filtro (Intersección)"], horizontal=True)

# Entrada de nuevo filtro
col1, col2 = st.columns(2)
with col1:
    clave = st.selectbox("🔑 Elige una categoría", sorted(FILTER_ENRICHED_OPTIONS.keys()), key=f"filtro_key_{len(st.session_state.filtros_aplicados)}")
with col2:
    # Calcular el número de especies para cada valor según el modo y filtros aplicados
    valores = sorted(FILTER_ENRICHED_OPTIONS[clave].keys())
    if modo.startswith("🔗") and st.session_state.especies_totales:
        especies_previas = set(st.session_state.especies_totales)
        valores_conteo = []
        for v in valores:
            especies_valor = set(FILTER_ENRICHED_OPTIONS[clave][v]["species"])
            interseccion = especies_previas & especies_valor
            valores_conteo.append((v, len(interseccion)))
        valor = st.selectbox(
            "🧬 Elige un valor",
            valores,
            format_func=lambda v: f"{v} ({dict(valores_conteo)[v]} especies)",
            key=f"filtro_val_{len(st.session_state.filtros_aplicados)}"
        )
    elif modo.startswith("➕"):
        valores_conteo = [(v, FILTER_ENRICHED_OPTIONS[clave][v]["count"]) for v in valores]
        valor = st.selectbox(
            "🧬 Elige un valor",
            valores,
            format_func=lambda v: f"{v} ({dict(valores_conteo)[v]} especies)",
            key=f"filtro_val_{len(st.session_state.filtros_aplicados)}"
        )
    else:
        valores_conteo = [(v, FILTER_ENRICHED_OPTIONS[clave][v]["count"]) for v in valores]
        valor = st.selectbox(
            "🧬 Elige un valor",
            valores,
            format_func=lambda v: f"{v} ({dict(valores_conteo)[v]} especies)",
            key=f"filtro_val_{len(st.session_state.filtros_aplicados)}"
        )

# Botón para añadir o combinar
if st.button("✅ Aplicar este filtro"):
    nuevo_filtro = {"tipo": "add" if modo.startswith("➕") else "combine", "clave": clave, "valor": valor}
    st.session_state.filtros_aplicados.append(nuevo_filtro)

    # Aplica el filtro individual
    filtros_dict = {clave: f"eq.{valor}"}
    especies = fetch_species_list(filters=filtros_dict)
    especies_nombres = [e["WithoutAutorship"] for e in especies]

    if nuevo_filtro["tipo"] == "add":
        st.session_state.especies_totales.extend(especies_nombres)
        st.success(f"Unión realizada.")
    elif nuevo_filtro["tipo"] == "combine":
        if st.session_state.especies_totales:
            especies_previas = set(st.session_state.especies_totales)
            especies_combinadas = [e for e in especies_nombres if e in especies_previas]
            st.session_state.especies_totales = especies_combinadas
            st.success("Combinación realizada.")
        else:
            st.session_state.especies_totales = especies_nombres
            st.success(f"Primera combinación aplicada.")

# Mostrar resumen de filtros aplicados
if st.session_state.filtros_aplicados:
    st.markdown("### 🧮 Filtros aplicados hasta ahora:")
    for i, f in enumerate(st.session_state.filtros_aplicados):
        st.markdown(f"- {i+1}. **{f['tipo'].capitalize()}**: {f['clave']} = {f['valor']}")

# Mostrar especies únicas y duplicadas
if st.session_state.especies_totales:
    todas = st.session_state.especies_totales
    únicas = sorted(set(todas))
    duplicadas = sorted({x for x in todas if todas.count(x) > 1})

    st.markdown(f"### 🧬 Total especies únicas seleccionadas: {len(únicas)}")
    st.markdown("Selecciona cuáles quieres conservar para la búsqueda final:")

    select_all = st.checkbox("Seleccionar todas", value=True)
    seleccionadas = st.multiselect("Especies", únicas, default=únicas if select_all else [])

    st.session_state.especies_seleccionadas_finales = seleccionadas

    if duplicadas:
        st.markdown(f"⚠️ **{len(duplicadas)} especies estaban duplicadas** entre los filtros. Solo se añadirán una vez:")
        for d in duplicadas:
            st.markdown(f"- 🔁 {d}")

# Lanzar búsqueda
if st.button("🔍 Ejecutar búsqueda de artículos"):
    if not st.session_state.especies_seleccionadas_finales:
        st.warning("Debes seleccionar al menos una especie.")
    else:
        # Usar el diccionario óptimo para pasar solo la lista de WithoutAutorship seleccionadas
        filtros_usados = {"_species": list(st.session_state.especies_seleccionadas_finales)}

        with st.spinner("Buscando artículos y procesando abstracts..."):
            df_resultado = update_species_articles(filters=filtros_usados, streamlit_mode=True)
            if df_resultado is not None:
                st.session_state["df_resultado"] = df_resultado

if "df_resultado" in st.session_state:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        st.session_state["df_resultado"].to_excel(writer, index=False)
    output.seek(0)

    st.download_button(
        "📥 Descargar Excel con artículos",
        data=output,
        file_name="ROSAL_IA.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.success("✅ Artículos encontrados y archivo generado, puedes descargarlo. n\\ Las variables del dataset son: `scientific name`, `title`, `year`, `authors`, `url`, `abstract`, `abs_pres` y `criterio`. abs_pres indica si el abstract está presente (1) o no (0). El criterio indica si el artículo es específico de la especie(`Exacto`) o no, pero es específico de otra especie del mismo género (`Genus`).")

# --- 3. Reporter: Informe en pdf ---
if "df_resultado" in st.session_state:
    st.markdown("---")
    st.markdown("## 🧾 Informe científico automatizado")
    st.markdown("""
    A continuación puedes generar un informe automatizado a partir de los artículos científicos asociados a las especies seleccionadas.

    - **Textos específicos**: basados en artículos con criterio `Exacto` y abstract.
    - **Textos genéricos**: basados en artículos con criterio `Genus` y abstract.
    - **Palabras clave**: extraídas automáticamente con NLP.
    """)

    df_resultado = st.session_state["df_resultado"]
    df_resultado['title'] = df_resultado['title'].fillna("").apply(clean_text)
    df_resultado['abstract'] = df_resultado['abstract'].fillna("").apply(clean_text)

    report_data = generate_scientific_report_data(df_resultado)
    
    if st.button("🧾 Generar Informe Final"):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        path = generate_pdf_report(report_data, st.session_state.filtros_aplicados, timestamp, df_resultado)
        with open(path, 'rb') as f:
            st.download_button("📥 Descargar PDF", f, file_name=path, mime="application/pdf")

# endregion